from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path
from typing import Any
import json
import re
import shutil
import uuid

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from . import models


APP_ROOT = Path(__file__).resolve().parents[2]
UPLOAD_DIR = APP_ROOT / "uploads"
EXPORT_DIR = APP_ROOT / "outputs" / "exports"
MONEY_ZERO = Decimal("0.00")
TWOPLACES = Decimal("0.01")

REQUIRED_FIELDS = ["SKU", "层级", "子物料编码", "子物料名称", "单位", "BOM用量", "是否末级"]
OPTIONAL_BOM_FIELDS = [
    "子物料规格",
    "子物料生产方式",
    "子项类型",
    "工段分类",
    "物料分类",
    "功能目的",
    "变更原因",
]


@dataclass
class ParsedRow:
    row_no: int
    data: dict[str, Any]


def money(value: Decimal | int | float | str | None) -> Decimal:
    if value is None or value == "":
        return MONEY_ZERO
    return Decimal(str(value)).quantize(TWOPLACES, rounding=ROUND_HALF_UP)


def dec(value: Any, default: Decimal | None = None) -> Decimal | None:
    if value is None:
        return default
    text = str(value).replace(",", "").strip()
    if text == "":
        return default
    if text.endswith("%"):
        text = str(Decimal(text[:-1].strip()) / Decimal("100"))
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return default


def clean_header(value: Any) -> str:
    return str(value or "").replace("\ufeff", "").replace("\n", "").strip().replace(" ", "")


def clean_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_currency(value: Any) -> str:
    text = clean_text(value).upper()
    if text in {"", "人民币", "RMB", "CNY", "¥"}:
        return "CNY"
    return text


def parse_bool_leaf(value: Any, level: str, next_level: str | None = None) -> bool:
    text = clean_text(value)
    if text in {"1", "是", "Y", "YES", "TRUE", "true", "末级"}:
        return True
    if text in {"0", "否", "N", "NO", "FALSE", "false", "非末级"}:
        return False
    if next_level:
        return not str(next_level).startswith(f"{level}.")
    return True


def parse_workbook(content: bytes, filename: str) -> tuple[list[str], list[ParsedRow], list[dict[str, Any]]]:
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise ValueError("请上传 .xlsx 或 .xlsm 底表文件")
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = detect_header_row(ws)
    headers = [clean_header(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
    warnings: list[dict[str, Any]] = []
    missing_headers = [field for field in REQUIRED_FIELDS if field not in headers]
    if missing_headers:
        raise ValueError(f"缺少必填表头：{', '.join(missing_headers)}")

    rows: list[ParsedRow] = []
    for row_no in range(header_row + 1, ws.max_row + 1):
        row = {headers[col - 1]: ws.cell(row_no, col).value for col in range(1, ws.max_column + 1) if headers[col - 1]}
        if not any(value not in (None, "") for value in row.values()):
            continue
        rows.append(ParsedRow(row_no=row_no, data=row))
    if not rows:
        raise ValueError("底表没有可解析的数据行")
    if header_row != 1:
        warnings.append({"level": "提示", "code": "DUAL_HEADER", "message": f"识别第 {header_row} 行为业务表头"})
    return headers, rows, warnings


def detect_header_row(ws) -> int:
    for row_no in range(1, min(ws.max_row, 5) + 1):
        headers = {clean_header(ws.cell(row_no, col).value) for col in range(1, ws.max_column + 1)}
        if {"SKU", "层级", "子物料编码"}.issubset(headers):
            return row_no
    return 1


def find_price_field(headers: list[str]) -> str | None:
    preferred = ["采购单价", "采购单价含税", "含税单价", "当前单价", "一供单价", "单价"]
    for field in preferred:
        if field in headers:
            return field
    candidates = [
        header
        for header in headers
        if "单价" in header and "委外" not in header and "加工" not in header and "物料成本" not in header
    ]
    return candidates[0] if candidates else None


def find_outsource_field(headers: list[str]) -> str | None:
    preferred = ["委外加工费", "委外加工单价", "加工费", "外协加工费"]
    for field in preferred:
        if field in headers:
            return field
    candidates = [header for header in headers if "委外" in header and "加工费" in header]
    return candidates[0] if candidates else None


def find_tax_field(headers: list[str]) -> str | None:
    for field in ["税率", "采购税率", "增值税率"]:
        if field in headers:
            return field
    return None


def find_currency_field(headers: list[str]) -> str | None:
    for field in ["币种", "货币"]:
        if field in headers:
            return field
    return None


def create_upload_version(
    db: Session,
    *,
    content: bytes,
    filename: str,
    sku: str | None,
    version_name: str | None,
    uploader: str = "system",
) -> dict[str, Any]:
    headers, parsed_rows, warnings = parse_workbook(content, filename)
    price_field = find_price_field(headers)
    outsource_field = find_outsource_field(headers)
    tax_field = find_tax_field(headers)
    currency_field = find_currency_field(headers)

    levels = [clean_text(row.data.get("层级")) for row in parsed_rows]
    normalized_rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    risk_candidates: list[dict[str, Any]] = []
    sku_values: set[str] = set()
    seen_keys: set[tuple[str, str, str]] = set()

    for idx, parsed in enumerate(parsed_rows):
        row = parsed.data
        row_sku = clean_text(row.get("SKU"))
        level = clean_text(row.get("层级"))
        material_code = clean_text(row.get("子物料编码"))
        material_name = clean_text(row.get("子物料名称"))
        unit = clean_text(row.get("单位"))
        qty = dec(row.get("BOM用量"))
        next_level = levels[idx + 1] if idx + 1 < len(levels) else None
        is_leaf = parse_bool_leaf(row.get("是否末级"), level, next_level)
        stage = clean_text(row.get("工段分类")) or None
        purchase_price = dec(row.get(price_field)) if price_field else None
        tax_rate = dec(row.get(tax_field), Decimal("0.13")) if tax_field else Decimal("0.13")
        currency = normalize_currency(row.get(currency_field)) if currency_field else "CNY"
        outsource_fee = dec(row.get(outsource_field), MONEY_ZERO) if outsource_field else MONEY_ZERO

        if row_sku:
            sku_values.add(row_sku)
        for field in REQUIRED_FIELDS:
            if clean_text(row.get(field)) == "":
                errors.append(_issue("阻断", parsed.row_no, field, "必填字段为空", "补充后重新上传"))
        if qty is None or qty <= 0:
            errors.append(_issue("阻断", parsed.row_no, "BOM用量", "BOM用量为空、非数字或 <= 0", "修正为大于 0 的数字"))
        if is_leaf and not stage:
            errors.append(_issue("阻断", parsed.row_no, "工段分类", "末级物料工段分类为空", "补充工段分类"))
        if currency != "CNY":
            errors.append(_issue("阻断", parsed.row_no, "币种", "第一版仅支持人民币 CNY", "改为 CNY 或人民币"))
        key = (row_sku, level, material_code)
        if key in seen_keys:
            warnings.append(_issue("警告", parsed.row_no, "子物料编码", "同 SKU + 层级 + 子物料编码重复", "确认是否重复投料"))
        seen_keys.add(key)
        if is_leaf and purchase_price is None:
            risk_candidates.append(
                {
                    "row_no": parsed.row_no,
                    "sku": row_sku,
                    "material_code": material_code,
                    "material_name": material_name,
                    "stage": stage or "未分类",
                    "risk_type": "价格缺失",
                    "missing_field": price_field or "采购单价",
                    "status": "待补价",
                    "impact": "允许底表先入库，但阻断成本版本确认/发布",
                }
            )

        material_cost_in = money((qty or MONEY_ZERO) * purchase_price) if purchase_price is not None else MONEY_ZERO
        material_cost_ex = money(material_cost_in / (Decimal("1") + (tax_rate or MONEY_ZERO))) if purchase_price is not None else MONEY_ZERO
        outsourcing_cost = money((qty or MONEY_ZERO) * (outsource_fee or MONEY_ZERO))
        normalized_rows.append(
            {
                "row_no": parsed.row_no,
                "sku": row_sku,
                "level": level,
                "material_code": material_code,
                "material_name": material_name,
                "spec": clean_text(row.get("子物料规格")) or None,
                "production_mode": clean_text(row.get("子物料生产方式")) or None,
                "item_type": clean_text(row.get("子项类型")) or None,
                "stage": stage,
                "material_category": clean_text(row.get("物料分类")) or None,
                "function_purpose": clean_text(row.get("功能目的")) or None,
                "change_reason": clean_text(row.get("变更原因")) or None,
                "unit": unit,
                "quantity": qty or MONEY_ZERO,
                "is_leaf": 1 if is_leaf else 0,
                "purchase_price_tax_included": purchase_price,
                "tax_rate": tax_rate or Decimal("0.13"),
                "currency": currency,
                "outsourcing_fee": outsource_fee or MONEY_ZERO,
                "material_cost_tax_included": material_cost_in,
                "material_cost_tax_excluded": material_cost_ex,
                "outsourcing_cost": outsourcing_cost,
                "total_cost_tax_included": money(material_cost_in + outsourcing_cost),
                "total_cost_tax_excluded": money(material_cost_ex + outsourcing_cost),
                "price_status": "MISSING_PRICE" if is_leaf and purchase_price is None else "OK",
                "raw_data": {key: _json_safe(value) for key, value in row.items()},
            }
        )

    if len(sku_values) != 1:
        errors.append(_issue("阻断", 0, "SKU", f"一个文件只能包含一个 SKU，当前识别到 {len(sku_values)} 个", "拆分为单 SKU 文件上传"))
    parsed_sku = next(iter(sku_values), "")
    if sku and parsed_sku and sku != parsed_sku:
        errors.append(_issue("阻断", 0, "SKU", f"页面 SKU {sku} 与底表 SKU {parsed_sku} 不一致", "保持页面与底表 SKU 一致"))

    if errors:
        return {
            "persisted": False,
            "errors": errors,
            "warnings": warnings,
            "summary": {
                "sku": parsed_sku or sku,
                "row_count": len(normalized_rows),
                "leaf_count": sum(1 for row in normalized_rows if row["is_leaf"]),
                "missing_price_count": len(risk_candidates),
                "price_field": price_field,
                "outsource_field": outsource_field,
            },
        }

    resolved_sku = sku or parsed_sku
    resolved_version_name = version_name or f"{resolved_sku}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    status = "NEEDS_PRICE" if risk_candidates else "CALCULATED"
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    stored_filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:8]}-{Path(filename).name}"
    file_path = UPLOAD_DIR / stored_filename
    file_path.write_bytes(content)

    batch = models.UploadBatch(
        sku=resolved_sku,
        filename=filename,
        file_path=str(file_path),
        uploader=uploader,
        row_count=len(normalized_rows),
        leaf_count=sum(1 for row in normalized_rows if row["is_leaf"]),
        missing_price_count=len(risk_candidates),
        status=status,
        parse_summary={
            "price_field": price_field,
            "outsource_field": outsource_field,
            "tax_field": tax_field,
            "currency_field": currency_field,
            "warnings": warnings,
        },
    )
    db.add(batch)
    db.flush()
    version = models.CostVersion(
        batch_id=batch.batch_id,
        sku=resolved_sku,
        version_name=resolved_version_name,
        status=status,
        missing_price_count=len(risk_candidates),
        total_tax_included=sum((row["total_cost_tax_included"] for row in normalized_rows if row["is_leaf"]), MONEY_ZERO),
        total_tax_excluded=sum((row["total_cost_tax_excluded"] for row in normalized_rows if row["is_leaf"]), MONEY_ZERO),
        created_by=uploader,
    )
    db.add(version)
    db.flush()

    for row in normalized_rows:
        db.add(models.BomItemSnapshot(version_id=version.version_id, **row))
    for risk in risk_candidates:
        db.add(models.RiskItem(version_id=version.version_id, **risk))
    db.commit()
    create_version_diff(db, version.version_id)
    db.refresh(version)
    return {
        "persisted": True,
        "batch": batch_to_dict(batch),
        "version": version_to_dict(version),
        "warnings": warnings,
        "errors": [],
        "summary": dashboard_summary(db, version.version_id, "tax_included"),
    }


def _issue(level: str, row_no: int, field: str, problem: str, suggestion: str) -> dict[str, Any]:
    return {"level": level, "row_no": row_no, "field": field, "problem": problem, "suggestion": suggestion}


def _json_safe(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def version_to_dict(version: models.CostVersion) -> dict[str, Any]:
    return {
        "version_id": version.version_id,
        "batch_id": version.batch_id,
        "sku": version.sku,
        "version_name": version.version_name,
        "status": version.status,
        "cost_basis": version.cost_basis,
        "total_tax_included": float(money(version.total_tax_included)),
        "total_tax_excluded": float(money(version.total_tax_excluded)),
        "missing_price_count": version.missing_price_count,
        "created_by": version.created_by,
        "created_at": version.created_at.isoformat() if version.created_at else None,
        "confirmed_at": version.confirmed_at.isoformat() if version.confirmed_at else None,
    }


def batch_to_dict(batch: models.UploadBatch) -> dict[str, Any]:
    return {
        "batch_id": batch.batch_id,
        "sku": batch.sku,
        "filename": batch.filename,
        "uploader": batch.uploader,
        "row_count": batch.row_count,
        "leaf_count": batch.leaf_count,
        "missing_price_count": batch.missing_price_count,
        "status": batch.status,
        "parse_summary": batch.parse_summary,
        "created_at": batch.created_at.isoformat() if batch.created_at else None,
    }


def item_to_dict(item: models.BomItemSnapshot) -> dict[str, Any]:
    return {
        "item_id": item.item_id,
        "row_no": item.row_no,
        "sku": item.sku,
        "level": item.level,
        "material_code": item.material_code,
        "material_name": item.material_name,
        "spec": item.spec,
        "production_mode": item.production_mode,
        "item_type": item.item_type,
        "stage": item.stage,
        "material_category": item.material_category,
        "function_purpose": item.function_purpose,
        "change_reason": item.change_reason,
        "unit": item.unit,
        "quantity": float(item.quantity),
        "is_leaf": bool(item.is_leaf),
        "purchase_price_tax_included": float(item.purchase_price_tax_included) if item.purchase_price_tax_included is not None else None,
        "tax_rate": float(item.tax_rate),
        "currency": item.currency,
        "outsourcing_fee": float(money(item.outsourcing_fee)),
        "material_cost_tax_included": float(money(item.material_cost_tax_included)),
        "material_cost_tax_excluded": float(money(item.material_cost_tax_excluded)),
        "outsourcing_cost": float(money(item.outsourcing_cost)),
        "total_cost_tax_included": float(money(item.total_cost_tax_included)),
        "total_cost_tax_excluded": float(money(item.total_cost_tax_excluded)),
        "price_status": item.price_status,
    }


def risk_to_dict(risk: models.RiskItem) -> dict[str, Any]:
    return {
        "risk_id": risk.risk_id,
        "sku": risk.sku,
        "row_no": risk.row_no,
        "material_code": risk.material_code,
        "material_name": risk.material_name,
        "stage": risk.stage,
        "risk_type": risk.risk_type,
        "missing_field": risk.missing_field,
        "status": risk.status,
        "impact": risk.impact,
    }


def diff_to_dict(diff: models.VersionDiff) -> dict[str, Any]:
    return {
        "diff_id": diff.diff_id,
        "current_version_id": diff.current_version_id,
        "previous_version_id": diff.previous_version_id,
        "sku": diff.sku,
        "material_code": diff.material_code,
        "material_name": diff.material_name,
        "diff_type": diff.diff_type,
        "previous_cost": float(money(diff.previous_cost)),
        "current_cost": float(money(diff.current_cost)),
        "diff_amount": float(money(diff.diff_amount)),
        "diff_ratio": float(diff.diff_ratio) if diff.diff_ratio is not None else None,
        "auto_reason": diff.auto_reason,
        "manual_reason": diff.manual_reason,
        "edited_by": diff.edited_by,
        "edited_at": diff.edited_at.isoformat() if diff.edited_at else None,
    }


def list_versions(db: Session) -> list[dict[str, Any]]:
    versions = db.scalars(select(models.CostVersion).order_by(models.CostVersion.created_at.desc())).all()
    return [version_to_dict(version) for version in versions]


def get_version_or_raise(db: Session, version_id: str) -> models.CostVersion:
    version = db.get(models.CostVersion, version_id)
    if not version:
        raise ValueError("成本版本不存在")
    return version


def get_version_detail(db: Session, version_id: str) -> dict[str, Any]:
    version = get_version_or_raise(db, version_id)
    items = db.scalars(
        select(models.BomItemSnapshot).where(models.BomItemSnapshot.version_id == version_id).order_by(models.BomItemSnapshot.row_no)
    ).all()
    risks = db.scalars(select(models.RiskItem).where(models.RiskItem.version_id == version_id).order_by(models.RiskItem.row_no)).all()
    return {
        "version": version_to_dict(version),
        "items": [item_to_dict(item) for item in items],
        "risks": [risk_to_dict(risk) for risk in risks],
    }


def delete_version(db: Session, version_id: str) -> dict[str, Any]:
    version = get_version_or_raise(db, version_id)
    result = version_to_dict(version)
    db.execute(
        delete(models.VersionDiff).where(
            (models.VersionDiff.current_version_id == version_id) | (models.VersionDiff.previous_version_id == version_id)
        )
    )
    db.delete(version)
    db.commit()
    return result


def recalculate_version(db: Session, version_id: str) -> dict[str, Any]:
    version = get_version_or_raise(db, version_id)
    items = db.scalars(select(models.BomItemSnapshot).where(models.BomItemSnapshot.version_id == version_id)).all()
    missing = 0
    for item in items:
        if not item.is_leaf:
            continue
        if item.purchase_price_tax_included is None:
            missing += 1
            item.price_status = "MISSING_PRICE"
            item.material_cost_tax_included = MONEY_ZERO
            item.material_cost_tax_excluded = MONEY_ZERO
        else:
            item.price_status = "OK"
            item.material_cost_tax_included = money(item.quantity * item.purchase_price_tax_included)
            item.material_cost_tax_excluded = money(item.material_cost_tax_included / (Decimal("1") + item.tax_rate))
        item.outsourcing_cost = money(item.quantity * item.outsourcing_fee)
        item.total_cost_tax_included = money(item.material_cost_tax_included + item.outsourcing_cost)
        item.total_cost_tax_excluded = money(item.material_cost_tax_excluded + item.outsourcing_cost)
    leaf_items = [item for item in items if item.is_leaf]
    version.total_tax_included = sum((money(item.total_cost_tax_included) for item in leaf_items), MONEY_ZERO)
    version.total_tax_excluded = sum((money(item.total_cost_tax_excluded) for item in leaf_items), MONEY_ZERO)
    version.missing_price_count = missing
    version.status = "NEEDS_PRICE" if missing else "CALCULATED"
    db.commit()
    create_version_diff(db, version_id)
    db.refresh(version)
    return version_to_dict(version)


def confirm_version(db: Session, version_id: str) -> dict[str, Any]:
    version = get_version_or_raise(db, version_id)
    if version.missing_price_count > 0:
        raise ValueError("存在缺失价格物料，不能确认/发布成本版本")
    version.status = "CONFIRMED"
    version.confirmed_at = datetime.utcnow()
    db.commit()
    db.refresh(version)
    return version_to_dict(version)


def previous_version_for(db: Session, current: models.CostVersion) -> models.CostVersion | None:
    return db.scalars(
        select(models.CostVersion)
        .where(models.CostVersion.sku == current.sku, models.CostVersion.created_at < current.created_at)
        .order_by(models.CostVersion.created_at.desc())
    ).first()


def create_version_diff(db: Session, current_version_id: str, previous_version_id: str | None = None) -> list[dict[str, Any]]:
    current = get_version_or_raise(db, current_version_id)
    previous = db.get(models.CostVersion, previous_version_id) if previous_version_id else previous_version_for(db, current)
    db.execute(delete(models.VersionDiff).where(models.VersionDiff.current_version_id == current_version_id))
    if previous is None:
        db.commit()
        return []

    current_items = {
        item.material_code: item
        for item in db.scalars(
            select(models.BomItemSnapshot).where(models.BomItemSnapshot.version_id == current.version_id, models.BomItemSnapshot.is_leaf == 1)
        ).all()
    }
    previous_items = {
        item.material_code: item
        for item in db.scalars(
            select(models.BomItemSnapshot).where(models.BomItemSnapshot.version_id == previous.version_id, models.BomItemSnapshot.is_leaf == 1)
        ).all()
    }
    diffs: list[models.VersionDiff] = []
    for material_code in sorted(set(current_items) | set(previous_items)):
        cur = current_items.get(material_code)
        prev = previous_items.get(material_code)
        types: list[str] = []
        if cur and not prev:
            types.append("新增物料")
        elif prev and not cur:
            types.append("删除物料")
        else:
            assert cur is not None and prev is not None
            if money(cur.quantity) != money(prev.quantity):
                types.append("用量变化")
            if money(cur.purchase_price_tax_included) != money(prev.purchase_price_tax_included):
                types.append("单价变化")
            if money(cur.outsourcing_fee) != money(prev.outsourcing_fee):
                types.append("委外加工费变化")
            if (cur.stage or "") != (prev.stage or ""):
                types.append("工段变化")
            if (cur.spec or "") != (prev.spec or ""):
                types.append("规格变化")
        cur_cost = money(cur.total_cost_tax_included if cur else 0)
        prev_cost = money(prev.total_cost_tax_included if prev else 0)
        amount = money(cur_cost - prev_cost)
        if not types and amount == MONEY_ZERO:
            continue
        ratio = None if prev_cost == MONEY_ZERO else (amount / prev_cost).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
        source = cur or prev
        auto_reason = getattr(cur, "change_reason", None) or "、".join(types)
        diffs.append(
            models.VersionDiff(
                current_version_id=current.version_id,
                previous_version_id=previous.version_id,
                sku=current.sku,
                material_code=material_code,
                material_name=source.material_name,
                diff_type="、".join(types) or "成本变化",
                previous_cost=prev_cost,
                current_cost=cur_cost,
                diff_amount=amount,
                diff_ratio=ratio,
                auto_reason=auto_reason,
            )
        )
    db.add_all(diffs)
    db.commit()
    return [diff_to_dict(diff) for diff in diffs]


def list_diffs(db: Session, version_id: str, previous_version_id: str | None = None) -> list[dict[str, Any]]:
    if previous_version_id:
        create_version_diff(db, version_id, previous_version_id)
    diffs = db.scalars(
        select(models.VersionDiff).where(models.VersionDiff.current_version_id == version_id).order_by(models.VersionDiff.diff_amount.desc())
    ).all()
    return [diff_to_dict(diff) for diff in diffs]


def update_diff_reason(db: Session, diff_id: str, manual_reason: str, edited_by: str = "system") -> dict[str, Any]:
    diff = db.get(models.VersionDiff, diff_id)
    if not diff:
        raise ValueError("差异记录不存在")
    diff.manual_reason = manual_reason
    diff.edited_by = edited_by
    diff.edited_at = datetime.utcnow()
    db.commit()
    db.refresh(diff)
    return diff_to_dict(diff)


def dashboard_summary(db: Session, version_id: str | None = None, cost_basis: str = "tax_included") -> dict[str, Any]:
    version = get_version_or_raise(db, version_id) if version_id else db.scalars(select(models.CostVersion).order_by(models.CostVersion.created_at.desc())).first()
    if version is None:
        return empty_dashboard()
    items = db.scalars(
        select(models.BomItemSnapshot).where(models.BomItemSnapshot.version_id == version.version_id, models.BomItemSnapshot.is_leaf == 1)
    ).all()
    cost_attr = "total_cost_tax_excluded" if cost_basis == "tax_excluded" else "total_cost_tax_included"
    total = money(version.total_tax_excluded if cost_basis == "tax_excluded" else version.total_tax_included)
    stage_map: dict[str, Decimal] = defaultdict(lambda: MONEY_ZERO)
    missing_by_stage: dict[str, int] = defaultdict(int)
    for item in items:
        stage = item.stage or "未分类"
        stage_map[stage] += money(getattr(item, cost_attr))
        if item.price_status == "MISSING_PRICE":
            missing_by_stage[stage] += 1
    stage_summary = [
        {
            "stage": stage,
            "cost": float(money(cost)),
            "ratio": float((money(cost) / total).quantize(Decimal("0.0001"))) if total else 0,
            "missing_price_count": missing_by_stage.get(stage, 0),
        }
        for stage, cost in sorted(stage_map.items(), key=lambda x: x[1], reverse=True)
    ]
    diffs = list_diffs(db, version.version_id)
    diff_amount = sum((Decimal(str(diff["diff_amount"])) for diff in diffs), MONEY_ZERO)
    previous_total = total - diff_amount
    risks = db.scalars(select(models.RiskItem).where(models.RiskItem.version_id == version.version_id).order_by(models.RiskItem.stage)).all()
    trend_versions = db.scalars(
        select(models.CostVersion).where(models.CostVersion.sku == version.sku).order_by(models.CostVersion.created_at)
    ).all()
    return {
        "version": version_to_dict(version),
        "cost_basis": cost_basis,
        "kpis": {
            "total_cost": float(total),
            "diff_amount": float(money(diff_amount)),
            "diff_ratio": float((money(diff_amount) / previous_total).quantize(Decimal("0.0001"))) if previous_total else None,
            "max_cost_stage": stage_summary[0]["stage"] if stage_summary else "-",
            "max_increase_stage": max(stage_summary, key=lambda row: row["cost"])["stage"] if stage_summary else "-",
            "missing_price_count": version.missing_price_count,
        },
        "stage_summary": stage_summary,
        "risks": [risk_to_dict(risk) for risk in risks],
        "trend": [
            {
                "version_id": item.version_id,
                "version_name": item.version_name,
                "cost": float(money(item.total_tax_excluded if cost_basis == "tax_excluded" else item.total_tax_included)),
                "status": item.status,
            }
            for item in trend_versions
        ],
        "diffs": diffs[:20],
    }


def empty_dashboard() -> dict[str, Any]:
    return {
        "version": None,
        "cost_basis": "tax_included",
        "kpis": {
            "total_cost": 0,
            "diff_amount": 0,
            "diff_ratio": None,
            "max_cost_stage": "-",
            "max_increase_stage": "-",
            "missing_price_count": 0,
        },
        "stage_summary": [],
        "risks": [],
        "trend": [],
        "diffs": [],
    }


def create_template_workbook() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM底表上传"
    owner = ["财务公式"] + ["研发填写"] * 13 + ["财务填写"] * 4
    headers = REQUIRED_FIELDS[:1] + ["层级", "子物料编码", "子物料名称"] + OPTIONAL_BOM_FIELDS + ["单位", "BOM用量", "是否末级", "采购单价", "税率", "币种", "委外加工费"]
    ws.append(owner[: len(headers)])
    ws.append(headers)
    ws.append(["SKU-001", "0", "SKU-001", "示例成品", "成品规格", "委外", "标准件", "整机", "成品", "", "", "PCS", 1, "否", "", "13%", "CNY", ""])
    ws.append(["SKU-001", "1", "MAT-001", "示例末级物料", "规格A", "外购", "标准件", "PCBA", "电子件", "", "初版", "PCS", 2, "是", 12.34, "13%", "CNY", 0])
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="1F3A5F" if cell.row == 2 else "EAF3FF")
            cell.font = Font(color="FFFFFF" if cell.row == 2 else "1F3A5F", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def export_dashboard_workbook(db: Session, version_id: str, cost_basis: str = "tax_included") -> BytesIO:
    summary = dashboard_summary(db, version_id, cost_basis)
    detail = get_version_detail(db, version_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "经营看板"
    ws.append(["产品成本管控看板", "", "", ""])
    ws.append(["SKU", summary["version"]["sku"], "成本版本", summary["version"]["version_name"]])
    ws.append(["成本口径", "不含税" if cost_basis == "tax_excluded" else "含税", "版本状态", summary["version"]["status"]])
    for key, value in summary["kpis"].items():
        ws.append([key, value])
    ws2 = wb.create_sheet("工段成本")
    ws2.append(["工段", "成本", "占比", "缺价数"])
    for row in summary["stage_summary"]:
        ws2.append([row["stage"], row["cost"], row["ratio"], row["missing_price_count"]])
    ws3 = wb.create_sheet("缺价风险")
    ws3.append(["SKU", "物料编码", "物料名称", "所属工段", "缺失字段", "状态", "影响"])
    for risk in summary["risks"]:
        ws3.append([risk["sku"], risk["material_code"], risk["material_name"], risk["stage"], risk["missing_field"], risk["status"], risk["impact"]])
    ws4 = wb.create_sheet("成本明细")
    ws4.append(["行号", "层级", "物料编码", "物料名称", "工段", "用量", "含税单价", "税率", "委外费", "含税总成本", "不含税总成本", "价格状态"])
    for item in detail["items"]:
        if item["is_leaf"]:
            ws4.append([
                item["row_no"],
                item["level"],
                item["material_code"],
                item["material_name"],
                item["stage"],
                item["quantity"],
                item["purchase_price_tax_included"],
                item["tax_rate"],
                item["outsourcing_fee"],
                item["total_cost_tax_included"],
                item["total_cost_tax_excluded"],
                item["price_status"],
            ])
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 18
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
